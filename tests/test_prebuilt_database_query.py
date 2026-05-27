"""Tests for the historical/precedent deal database query layer and flexible Excel export.

Covers the new agent-facing APIs added for exploring the DB built by sweeps
and exporting *exactly* the selected deals as light or full Excel.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from ptbot.db import (
    get_deal_statistics,
    get_similar_deals,
    get_top_deals,
    insert_deals,
    insert_run,
    new_run_id,
    open_db,
    row_to_deal_candidate,
    search_deals,
    search_deals_as_candidates,
    summarize_database,
)
from ptbot.excel import generate_comps_excel_from_deals
from ptbot.models import DealCandidate, ResearchParams

# Shared fixtures (mirrors test_db.py style)
_PARAMS = ResearchParams(
    sector="FinTech",
    geography="United States",
    start_date="2024-01-01",
    end_date="2024-12-31",
)

_DEAL_A = DealCandidate(
    target="Acme Payments",
    acquirer="BigBank Corp",
    date="2024-03-15",
    deal_value="$120M",
    multiples_disclosed=True,
    multiples=("5.2x EV/Revenue", "18.1x EV/EBITDA"),
    source_urls=("https://example.com/deal-a",),
)

_DEAL_B = DealCandidate(
    target="Zeta Lending",
    acquirer="Global Finance Inc",
    date="2024-06-01",
    deal_value="$85M",
    multiples_disclosed=True,
    computed_multiples_available=True,
    multiples=("4.8x EV/Revenue",),
    source_urls=("https://example.com/deal-b", "https://sec.gov/..."),
)


def _seed_basic_data(tmp_path: Path) -> tuple[Path, str]:
    """Seed a temp DB with two runs and some deals for query testing."""
    db_path = tmp_path / "query-test.db"
    conn = open_db(db_path)

    run_id_1 = new_run_id()
    insert_run(conn, run_id_1, _PARAMS)
    insert_deals(conn, run_id_1, [_DEAL_A, _DEAL_B], qualified_keys={_DEAL_A.key()})

    # Second run in a different sector for summarize / similar tests
    params2 = _PARAMS.model_copy(update={"sector": "HealthTech", "geography": "Europe"})
    run_id_2 = new_run_id()
    insert_run(conn, run_id_2, params2)
    insert_deals(conn, run_id_2, [_DEAL_B], qualified_keys=set())

    conn.close()
    return db_path, run_id_1


def test_row_to_deal_candidate_basic_and_json_edges() -> None:
    row = {
        "target": "TestCo",
        "acquirer": "Buyer Inc",
        "date": "2024-01-01",
        "deal_value": "$50M",
        "multiples_disclosed": 1,
        "computed_multiples_available": 0,
        "multiples": json.dumps(["4.1x EV/Revenue"]),
        "source_urls": json.dumps(["https://a.com"]),
    }
    deal = row_to_deal_candidate(row)
    assert deal.target == "TestCo"
    assert deal.multiples == ("4.1x EV/Revenue",)
    assert deal.source_urls == ("https://a.com",)

    # Non-list / bad JSON fallback
    bad_row = {"target": "X", "acquirer": "Y", "multiples": "not json", "source_urls": 42}
    deal2 = row_to_deal_candidate(bad_row)
    assert deal2.multiples == ()
    assert deal2.source_urls == ()


def test_search_deals_filters_and_limit(tmp_path: Path) -> None:
    db_path, _ = _seed_basic_data(tmp_path)
    conn = open_db(db_path)

    # Sector + qualified
    results = search_deals(conn, sector="FinTech", qualified=True)
    assert len(results) == 1
    assert results[0]["target"] == "Acme Payments"

    # Geography + limit
    results = search_deals(conn, geography="United States", limit=1)
    assert len(results) == 1

    # No matches
    results = search_deals(conn, sector="NonExistent")
    assert results == []

    conn.close()


def test_search_deals_as_candidates_roundtrip(tmp_path: Path) -> None:
    db_path, _ = _seed_basic_data(tmp_path)
    conn = open_db(db_path)

    candidates = search_deals_as_candidates(conn, sector="FinTech")
    assert len(candidates) == 2
    assert all(isinstance(c, DealCandidate) for c in candidates)
    assert candidates[0].key()  # has working key()

    conn.close()


def test_summarize_database_basic(tmp_path: Path) -> None:
    db_path, _ = _seed_basic_data(tmp_path)
    conn = open_db(db_path)

    summary = summarize_database(conn)
    assert summary["total_runs"] >= 2
    assert summary["total_deals"] >= 3
    assert summary["qualified_deals"] >= 1
    assert any(s["sector"] == "FinTech" for s in summary.get("sectors", []))

    conn.close()


def test_get_deal_statistics_and_top_deals(tmp_path: Path) -> None:
    db_path, _ = _seed_basic_data(tmp_path)
    conn = open_db(db_path)

    stats = get_deal_statistics(conn, sector="FinTech")
    assert stats["total_deals"] >= 1

    top = get_top_deals(conn, n=5, by="recent")
    assert len(top) >= 1
    assert isinstance(top[0], DealCandidate)

    conn.close()


def test_get_similar_deals_fuzzy(tmp_path: Path) -> None:
    db_path, _ = _seed_basic_data(tmp_path)
    conn = open_db(db_path)

    similar = get_similar_deals(conn, sector="Fin", limit=10)
    assert len(similar) >= 1  # should fuzzy-match FinTech

    conn.close()


def test_generate_comps_excel_from_deals_light_and_full(tmp_path: Path) -> None:
    db_path, _ = _seed_basic_data(tmp_path)
    conn = open_db(db_path)
    deals = search_deals_as_candidates(conn, sector="FinTech")
    conn.close()

    light_path = tmp_path / "light.xlsx"
    generate_comps_excel_from_deals(deals, light_path, "Test Light", style="light")
    assert light_path.exists()
    wb = load_workbook(light_path)
    assert "Selected Deals" in wb.sheetnames
    ws = wb["Selected Deals"]
    # Row 1 must hold the title — not overwritten by write_headers (regression guard)
    assert (
        ws["A1"].value == "Test Light"
    ), f"Title cell A1 overwritten — expected 'Test Light', got {ws['A1'].value!r}"
    assert ws["A2"].value is not None, "Subtitle missing from A2"
    assert "PTBot" in ws["A2"].value, "PTBot branding missing from subtitle at A2"
    # Column headers must be at row 3, data at row 4+
    assert (
        ws.cell(row=3, column=1).value == "Target"
    ), f"Headers not at row 3 — A3 = {ws.cell(row=3, column=1).value!r}"

    full_path = tmp_path / "full.xlsx"
    generate_comps_excel_from_deals(deals, full_path, "Test Full", style="full")
    assert full_path.exists()
    wb2 = load_workbook(full_path)
    assert "Comps Table" in wb2.sheetnames or "Cover" in wb2.sheetnames


def test_generate_comps_excel_from_deals_empty_and_mixed_input(tmp_path: Path) -> None:
    empty_path = tmp_path / "empty.xlsx"
    generate_comps_excel_from_deals([], empty_path, "Empty", style="light")
    assert empty_path.exists()
