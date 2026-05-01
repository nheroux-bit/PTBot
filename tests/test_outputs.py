"""Tests for PDF and Excel output helpers."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from ptbot.excel import generate_comps_excel, parse_multiple_columns
from ptbot.models import DealCandidate
from ptbot.pdf import markdown_to_pdf, sanitize_text, strip_markdown


def test_sanitize_text_replaces_common_unicode() -> None:
    """PDF sanitization should replace characters unsupported by core fonts."""
    assert sanitize_text("A — B ≥ C") == "A -- B >= C"


def test_strip_markdown_removes_links_and_bold() -> None:
    """Markdown syntax should be simplified for PDF text."""
    assert strip_markdown("**Deal** [source](https://example.com)") == "Deal source"


def test_markdown_to_pdf_writes_file(tmp_path: Path) -> None:
    """Markdown reports should render to a non-empty PDF."""
    markdown = tmp_path / "report.md"
    pdf = tmp_path / "report.pdf"
    markdown.write_text("# Report\n\n## Summary\n\n- One item", encoding="utf-8")

    markdown_to_pdf(markdown, pdf, "Report")

    assert pdf.exists()
    assert pdf.stat().st_size > 0


def test_generate_comps_excel_writes_expected_sheets(tmp_path: Path) -> None:
    """Excel generator should create canonical comps and excluded tabs."""
    deals = tmp_path / "qualified_deals.json"
    workbook_path = tmp_path / "comps.xlsx"
    deals.write_text(
        json.dumps(
            [
                {
                    "target": "Performant Healthcare, Inc.",
                    "acquirer": "Machinify",
                    "date": "2025-08-01",
                    "deal_value": "$670M",
                    "multiples": ["EV_Q2_annualized_Revenue: 4.4x", "premium_to_90d_VWAP: 139%"],
                    "source_urls": ["https://example.com/performant"],
                },
                {
                    "target": "Performant Healthcare, Inc. (Plantation, FL)",
                    "acquirer": "Machinify",
                    "date": "2025-08-01",
                    "deal_value": "$670M",
                    "multiples": ["EV_Q2_annualized_EBITDA: 27.0x"],
                    "source_urls": ["https://example.com/performant-2"],
                },
                {
                    "target": "Urbint, Inc. (Miami Beach, FL)",
                    "acquirer": "Itron",
                    "date": "2025-10-06",
                    "deal_value": "$330.7M",
                    "multiples": ["goodwill_pct_of_purchase_price: 77%"],
                    "source_urls": ["https://example.com/urbint"],
                }
            ]
        ),
        encoding="utf-8",
    )

    generate_comps_excel(deals, workbook_path, "AI Comps")

    workbook = load_workbook(workbook_path)
    assert workbook.sheetnames == [
        "Cover",
        "Comps Table",
        "Excluded Qualitative",
        "Benchmarks",
        "Sources",
    ]
    assert workbook["Cover"]["B5"].value == 3
    assert workbook["Cover"]["B6"].value == 1
    assert workbook["Cover"]["B7"].value == 1
    assert workbook["Comps Table"]["G1"].value == "Revenue Basis"
    assert workbook["Comps Table"]["H1"].value == "EBITDA"
    assert workbook["Comps Table"]["I1"].value == "EBITDA Basis"
    assert workbook["Comps Table"]["J1"].value == "ARR"
    assert workbook["Comps Table"]["K1"].value == "ARR Basis"
    assert workbook["Comps Table"]["N1"].value == "EV/Revenue"
    assert workbook["Comps Table"]["O1"].value == "EV/EBITDA"
    assert workbook["Comps Table"]["P1"].value == "EV/ARR"
    assert workbook["Comps Table"]["A2"].value == "Performant Healthcare, Inc."
    assert workbook["Comps Table"]["E2"].value == 670.0
    assert round(workbook["Comps Table"]["F2"].value, 1) == 152.3
    assert workbook["Comps Table"]["G2"].value == "Q2 annualized / run-rate"
    assert round(workbook["Comps Table"]["H2"].value, 1) == 24.8
    assert workbook["Comps Table"]["I2"].value == "Q2 annualized / run-rate"
    assert workbook["Comps Table"]["N2"].value == '=IFERROR(E2/F2,"")'
    assert workbook["Comps Table"]["O2"].value == '=IFERROR(E2/H2,"")'
    assert workbook["Comps Table"]["R2"].value == 1.39
    assert workbook["Comps Table"].freeze_panes == "A2"
    assert workbook["Excluded Qualitative"]["A2"].value == "Urbint, Inc. (Miami Beach, FL)"
    assert "Excluded from stats" in workbook["Excluded Qualitative"]["L2"].value
    assert workbook["Comps Table"]["N4"].value == '=IFERROR(AVERAGE(N2:N2),"")'
    assert workbook["Benchmarks"]["B6"].value == "='Comps Table'!N4"
    assert workbook["Comps Table"]["T2"].hyperlink.target == "https://example.com/performant"
    assert workbook["Sources"]["C2"].hyperlink.target == "https://example.com/performant"


def test_generate_comps_excel_rejects_non_list_json(tmp_path: Path) -> None:
    """Qualified deals JSON must be a list."""
    deals = tmp_path / "qualified_deals.json"
    deals.write_text(json.dumps({"target": "bad"}), encoding="utf-8")
    with pytest.raises(ValueError, match="must be a list"):
        generate_comps_excel(deals, tmp_path / "comps.xlsx", "Bad")


def test_parse_multiple_columns_prefers_normalized_revenue_multiple() -> None:
    """Forward/normalized revenue multiples should beat distorted LTM figures."""
    deal = DealCandidate(
        target="DigitalBridge",
        acquirer="SoftBank",
        multiples=("EV_LTM_Revenue: 42.9x", "EV_NTM_Revenue: 10.2x"),
    )

    assert parse_multiple_columns(deal)["EV/Revenue"] == 10.2
    assert parse_multiple_columns(deal)["Revenue Basis"] == "NTM"
