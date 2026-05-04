"""Excel comps workbook generation for PTBot."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .models import DealCandidate
from .orchestrator import dedupe_deals, filter_qualified_deals

BLUE = "0070C0"
BLACK = "000000"
HYPERLINK_BLUE = "0563C1"
HEADER_FILL = "D9EAF7"
EXCLUDED_FILL = "FCE4D6"
SUMMARY_FILL = "E2F0D9"


def _extract_number(text: str) -> float | None:
    """Extract the first multiple/percentage number from text."""
    search_text = text.split(":", maxsplit=1)[-1]
    match = re.search(r"~?(-?\d+(?:\.\d+)?)\s*[x×%]?", search_text, re.IGNORECASE)
    if match is None:
        return None
    return float(match.group(1))


def parse_deal_value_to_mm(deal_value: str | None) -> float | None:
    """Parse deal value text into USD millions where possible."""
    if not deal_value:
        return None
    match = re.search(r"\$?\s*~?(\d+(?:\.\d+)?)\s*([bm])\b", deal_value, re.IGNORECASE)
    if match is None:
        return None
    value = float(match.group(1))
    unit = match.group(2).lower()
    return value * 1000 if unit == "b" else value


def parse_multiple_columns(deal: DealCandidate) -> dict[str, float | str | None]:
    """Parse standard multiple strings into workbook-ready numeric columns."""
    parsed: dict[str, float | str | None] = {
        "EV/Revenue": None,
        "EV/EBITDA": None,
        "EV/ARR": None,
        "P/E": None,
        "Premium": None,
    }
    bases: dict[str, str | None] = {
        "Revenue Basis": None,
        "EBITDA Basis": None,
        "ARR Basis": None,
    }
    priorities = dict.fromkeys(parsed, -1)
    for item in deal.multiples:
        text = item.lower()
        value = _extract_number(item)
        if value is None:
            continue
        if any(excluded in text for excluded in ("goodwill", "backlog", "pipeline", "patent")):
            continue
        if has_ev_metric(text, "revenue") or has_ev_metric(text, "sales"):
            priority = multiple_priority(text)
            if priority > priorities["EV/Revenue"]:
                parsed["EV/Revenue"] = value
                bases["Revenue Basis"] = multiple_basis(text)
                priorities["EV/Revenue"] = priority
        elif has_ev_metric(text, "ebitda"):
            priority = multiple_priority(text)
            if priority > priorities["EV/EBITDA"]:
                parsed["EV/EBITDA"] = value
                bases["EBITDA Basis"] = multiple_basis(text)
                priorities["EV/EBITDA"] = priority
        elif has_ev_metric(text, "arr"):
            priority = multiple_priority(text)
            if priority > priorities["EV/ARR"]:
                parsed["EV/ARR"] = value
                bases["ARR Basis"] = multiple_basis(text)
                priorities["EV/ARR"] = priority
        elif "p/e" in text or "price/earnings" in text:
            parsed["P/E"] = parsed["P/E"] or value
        elif "premium" in text:
            parsed["Premium"] = parsed["Premium"] or value / 100
    return {**parsed, **bases}


def has_ev_metric(text: str, metric: str) -> bool:
    """Return whether text describes an EV-based standard multiple for a metric."""
    return bool(re.search(rf"\bev\b.*\b{metric}\b", text.replace("_", " ")))


def multiple_priority(text: str) -> int:
    """Rank multiple bases so normalized/forward metrics beat distorted historical ones."""
    if "ntm" in text or "fy" in text or "fee" in text:
        return 3
    if "ttm" in text or re.search(r"\bq\d\b", text) or "run-rate" in text or "annualized" in text:
        return 2
    if "ltm" in text:
        return 1
    return 0


def multiple_basis(text: str) -> str:
    """Return a concise denominator basis label for a parsed multiple."""
    if "ntm" in text:
        return "NTM"
    if "ttm" in text:
        return "TTM"
    if "q2" in text or "run-rate" in text or "annualized" in text:
        return "Q2 annualized / run-rate"
    if "ltm" in text:
        return "LTM"
    if "fy" in text or "fee" in text:
        return "FY / fee revenue"
    return "Unspecified"


def caveat_for_deal(deal: DealCandidate, included: bool) -> str:
    """Return a concise workbook caveat for a deal."""
    text = " ".join((deal.target, deal.acquirer, deal.deal_value or "", *deal.multiples)).lower()
    caveats: list[str] = []
    if not included:
        caveats.append("Excluded from stats")
    if "loi" in text:
        caveats.append("LOI-stage")
    if "pending" in text or "expected close" in text:
        caveats.append("Pending close")
    if any(term in text for term in ("goodwill", "backlog", "pipeline", "patent", "licensing")):
        caveats.append("Qualitative/proxy metrics only")
    if "near-miss" in text or "rockville" in text or "maryland" in text:
        caveats.append("Geography near-miss")
    if not caveats:
        caveats.append("Included")
    return "; ".join(dict.fromkeys(caveats))


def load_deals(path: Path) -> list[DealCandidate]:
    """Load qualified deals from JSON."""
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError("qualified deals JSON must be a list")
    return [DealCandidate.model_validate(item) for item in data]


def prepare_deals(deals: list[DealCandidate]) -> tuple[list[DealCandidate], list[DealCandidate]]:
    """Deduplicate deals and split included comps from qualitative/excluded rows."""
    canonical = dedupe_deals(deals)
    included = filter_qualified_deals(canonical, 1)
    included_keys = {deal.key() for deal in included}
    excluded = [deal for deal in canonical if deal.key() not in included_keys]
    return included, excluded


def write_headers(sheet: Worksheet, headers: list[str], fill: str = HEADER_FILL) -> None:
    """Write formatted worksheet headers."""
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor=fill)
        cell.alignment = Alignment(wrap_text=True)


def write_deal_row(sheet: Worksheet, row: int, deal: DealCandidate, included: bool) -> None:
    """Write one deal row to a workbook sheet."""
    multiples = parse_multiple_columns(deal)
    enterprise_value = parse_deal_value_to_mm(deal.deal_value)
    ev_revenue = multiples["EV/Revenue"]
    ev_ebitda = multiples["EV/EBITDA"]
    ev_arr = multiples["EV/ARR"]
    revenue = implied_metric(
        enterprise_value, float(ev_revenue) if isinstance(ev_revenue, (int, float)) else None
    )
    ebitda = implied_metric(
        enterprise_value, float(ev_ebitda) if isinstance(ev_ebitda, (int, float)) else None
    )
    arr = implied_metric(
        enterprise_value, float(ev_arr) if isinstance(ev_arr, (int, float)) else None
    )
    values = [
        deal.target,
        deal.acquirer,
        deal.date or "",
        deal.deal_value or "",
        enterprise_value,
        revenue,
        multiples["Revenue Basis"],
        ebitda,
        multiples["EBITDA Basis"],
        arr,
        multiples["ARR Basis"],
        caveat_for_deal(deal, included),
        included,
        f'=IFERROR(E{row}/F{row},"")',
        f'=IFERROR(E{row}/H{row},"")',
        f'=IFERROR(E{row}/J{row},"")',
        multiples["P/E"],
        multiples["Premium"],
        "; ".join(deal.multiples),
        "; ".join(deal.source_urls),
    ]
    for col, value in enumerate(values, start=1):
        cell = sheet.cell(row=row, column=col, value=value)
        cell.font = Font(color=BLUE)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if col in {5, 6, 8, 10} and isinstance(value, int | float):
            cell.number_format = "$0.0"
        if col in {14, 15, 16, 17}:
            cell.number_format = "0.0x"
        if col == 18 and isinstance(value, int | float):
            cell.number_format = "0.0%"
    first_source = next(iter(deal.source_urls), None)
    if first_source:
        link_cell(sheet.cell(row=row, column=20), first_source)


def link_cell(cell: Cell, target: str) -> None:
    """Format a cell as a clickable hyperlink."""
    cell.hyperlink = target
    cell.style = "Hyperlink"
    cell.font = Font(color=HYPERLINK_BLUE, underline="single")


def implied_metric(enterprise_value: float | None, multiple: float | None) -> float | None:
    """Calculate an implied denominator from enterprise value and a multiple."""
    if enterprise_value is None or multiple is None or multiple == 0:
        return None
    return enterprise_value / multiple


def generate_comps_excel(qualified_deals_path: Path, output_path: Path, title: str) -> None:
    """Generate an IB-formatted precedent comps workbook."""
    raw_deals = load_deals(qualified_deals_path)
    included_deals, excluded_deals = prepare_deals(raw_deals)
    workbook = Workbook()
    cover = workbook.active
    cover.title = "Cover"
    cover["A1"] = title
    cover["A1"].font = Font(bold=True, size=16)
    cover["A3"] = "Generated by PTBot"
    cover["A5"] = "Raw deal rows"
    cover["B5"] = len(raw_deals)
    cover["B5"].font = Font(color=BLUE)
    cover["A6"] = "Included comps"
    cover["B6"] = len(included_deals)
    cover["B6"].font = Font(color=BLUE)
    cover["A7"] = "Excluded / qualitative"
    cover["B7"] = len(excluded_deals)
    cover["B7"].font = Font(color=BLUE)

    comps = workbook.create_sheet("Comps Table")
    headers = [
        "Target",
        "Acquirer",
        "Date",
        "Deal Value Text",
        "Enterprise Value",
        "Revenue",
        "Revenue Basis",
        "EBITDA",
        "EBITDA Basis",
        "ARR",
        "ARR Basis",
        "Caveat",
        "Included in Stats?",
        "EV/Revenue",
        "EV/EBITDA",
        "EV/ARR",
        "P/E",
        "Premium",
        "Multiples Text",
        "Sources",
    ]
    write_headers(comps, headers)
    for row, deal in enumerate(included_deals, start=2):
        write_deal_row(comps, row, deal, included=True)

    summary_row = len(included_deals) + 3
    comps.cell(row=summary_row, column=1, value="Mean")
    comps.cell(row=summary_row, column=1).fill = PatternFill(
        fill_type="solid", fgColor=SUMMARY_FILL
    )
    last_data_row = max(2, len(included_deals) + 1)
    for col in range(14, 19):
        formula = f'=IFERROR(AVERAGE({cell_ref(col)}2:{cell_ref(col)}{last_data_row}),"")'
        cell = comps.cell(row=summary_row, column=col, value=formula)
        cell.font = Font(color=BLACK, bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor=SUMMARY_FILL)
        cell.number_format = "0.0%" if col == 18 else "0.0x"
    comps.freeze_panes = "A2"
    comps.auto_filter.ref = f"A1:T{max(1, len(included_deals) + 1)}"

    excluded = workbook.create_sheet("Excluded Qualitative")
    write_headers(excluded, headers, EXCLUDED_FILL)
    for row, deal in enumerate(excluded_deals, start=2):
        write_deal_row(excluded, row, deal, included=False)
    excluded.freeze_panes = "A2"
    excluded.auto_filter.ref = f"A1:T{max(1, len(excluded_deals) + 1)}"

    benchmarks = workbook.create_sheet("Benchmarks")
    benchmarks["A1"] = "Benchmark Context"
    benchmarks["A1"].font = Font(bold=True)
    benchmarks["A3"] = (
        "Use sector-wide benchmarks from the final deliverable to supplement local comps."
    )
    benchmarks["A5"] = "Metric"
    benchmarks["B5"] = "Local Mean"
    benchmarks["A6"] = "EV/Revenue"
    benchmarks["B6"] = f"='Comps Table'!N{summary_row}"
    benchmarks["A7"] = "EV/EBITDA"
    benchmarks["B7"] = f"='Comps Table'!O{summary_row}"
    benchmarks["A8"] = "EV/ARR"
    benchmarks["B8"] = f"='Comps Table'!P{summary_row}"
    benchmarks["A9"] = "Premium"
    benchmarks["B9"] = f"='Comps Table'!R{summary_row}"
    for row in range(5, 10):
        benchmarks.cell(row=row, column=1).font = Font(bold=row == 5)
        benchmarks.cell(row=row, column=2).font = Font(color=BLACK, bold=row == 5)

    sources = workbook.create_sheet("Sources")
    write_headers(sources, ["Target", "Included in Stats?", "Source URL"])
    source_row = 2
    included_keys = {item.key() for item in included_deals}
    for deal in [*included_deals, *excluded_deals]:
        included = deal.key() in included_keys
        for source in deal.source_urls:
            sources.cell(row=source_row, column=1, value=deal.target)
            sources.cell(row=source_row, column=2, value=included)
            sources.cell(row=source_row, column=3, value=source)
            link_cell(sources.cell(row=source_row, column=3), source)
            source_row += 1
    sources.freeze_panes = "A2"
    sources.auto_filter.ref = f"A1:C{max(1, source_row - 1)}"

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells) + 2
            sheet.column_dimensions[column_cells[0].column_letter].width = min(width, 60)

    workbook.save(output_path)


def cell_ref(col: int) -> str:
    """Return Excel column letters for a 1-indexed column number."""
    letters = ""
    while col:
        col, remainder = divmod(col - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters
